from datetime import date
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.core.database import Base
from app.models.models import (
    Preliquidacion, PreliquidacionLinea, ConceptoAdicional, TipoConcepto,
)
from app.services.export_service import COLUMNAS, generar_export_excel


@pytest.fixture()
def db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()


def _preliq(db):
    p = Preliquidacion(quincena=date(2026, 5, 1), creado_por=1)
    db.add(p)
    db.commit()
    db.refresh(p)
    return p


def _linea(db, preliq, **kwargs):
    l = PreliquidacionLinea(
        preliquidacion_id=preliq.id,
        nombre_tarea="PODA", nombre_cliente="CLI", nombre_finca="FINCA",
        importe_total=Decimal("0"), linea_incompleta=False,
        **kwargs,
    )
    db.add(l)
    db.commit()
    db.refresh(l)
    return l


def _hoja(buffer):
    return load_workbook(buffer, read_only=True)["Preliquidacion"]


def test_header_con_columnas_de_linea():
    assert COLUMNAS == [
        "Empresa", "planilla", "fecha_tarea", "nombre_cliente", "nombre_finca",
        "nombre_tarea", "nombre_tractor", "legajo", "nombre_empleado", "supervisor",
        "hsjornal", "hsmaquina", "unidades", "tancadas",
        "codigo", "cantidad", "precio", "importe", "grupo_pago", "duplicado",
    ]


def test_fila_emite_supervisor_y_cantidades(db):
    preliq = _preliq(db)
    linea = _linea(
        db, preliq,
        nombre_empleado="PEREZ JUAN", nombre_supervisor="GOMEZ ANA",
        hsjornal=Decimal("7.5"), hsmaquina=Decimal("2"),
        unidades=Decimal("300"), tancadas=Decimal("4"),
    )
    db.add(ConceptoAdicional(
        linea_id=linea.id, descripcion="X", codigo_concepto=11,
        tipo=TipoConcepto.OTRO, importe=Decimal("100"), ingresado_por=1,
    ))
    db.commit()

    ws = _hoja(generar_export_excel(db, preliq.id))
    filas = list(ws.iter_rows(values_only=True))
    header, fila = filas[0], filas[1]
    assert list(header) == COLUMNAS

    valores = dict(zip(header, fila))
    assert valores["supervisor"] == "GOMEZ ANA"
    assert valores["hsjornal"] == 7.5
    assert valores["hsmaquina"] == 2
    assert valores["unidades"] == 300
    assert valores["tancadas"] == 4
    assert valores["codigo"] == 11
    assert valores["importe"] == 100


def test_linea_sin_conceptos_igual_sale_con_cantidades(db):
    preliq = _preliq(db)
    _linea(db, preliq, nombre_supervisor=None, hsjornal=Decimal("6"))

    ws = _hoja(generar_export_excel(db, preliq.id))
    filas = list(ws.iter_rows(values_only=True))
    valores = dict(zip(filas[0], filas[1]))
    assert valores["hsjornal"] == 6
    assert valores["supervisor"] is None
    assert valores["codigo"] is None
